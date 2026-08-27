#!/usr/bin/env python3
"""Build MASTER_INVENTORY.xlsx from the CSV data files in ../data.

Pricing-tier logic (documented on the Assumptions sheet):
  Slabs:      Quick = List A (Jan-2025 offer sheet) x 0.70  | Mid = List A | Slow = List B (Jun-2025 list; falls back to List A)
  Equipment:  Quick = 55% of unit cost | Mid = 75% of cost  | Slow = sales list price (or 95% of cost if none)
  Appliances: Quick = 45% of MSRP      | Mid = 55% of MSRP  | Slow = 70% of MSRP
All tier cells are formulas so the sheet recalculates when inputs change.
"""
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
OUT = ROOT / "output" / "MASTER_INVENTORY.xlsx"

ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
BASE_FONT = Font(name=ARIAL, size=10)
BOLD = Font(name=ARIAL, bold=True, size=10)
TITLE_FONT = Font(name=ARIAL, bold=True, size=13)
MONEY = "$#,##0"
MONEY2 = "$#,##0.00"


def read_csv(name):
    with open(DATA / name, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    return rows[0], rows[1:]


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def num(v):
    if v is None or v == "":
        return None
    try:
        return float(v) if "." in v else int(v)
    except ValueError:
        return v


wb = Workbook()

# ---------------------------------------------------------------- Slabs
ws = wb.active
ws.title = "Slabs"
ws["A1"] = "PORCELAIN / QUARTZ / MARBLE SLAB INVENTORY (BZ International Trade LLC records)"
ws["A1"].font = TITLE_FONT
hdr, rows = read_csv("slabs_inventory.csv")
headers = hdr + ["Quick Sale $/pc", "Mid $/pc", "Slow $/pc", "Total @ Quick", "Total @ Mid", "Total @ Slow"]
ws.append([])
ws.append(headers)
style_header(ws, 3, len(headers))
first = 4
for r in rows:
    ws.append([r[0], r[1], r[2], r[3], num(r[4]), num(r[5]), num(r[6]), num(r[7]), r[8]])
last = ws.max_row
for i in range(first, last + 1):
    ws[f"J{i}"] = f"=IF(F{i}<>\"\",F{i}*Assumptions!$B$4,\"\")"
    ws[f"K{i}"] = f"=IF(F{i}<>\"\",F{i},\"\")"
    ws[f"L{i}"] = f"=IF(G{i}<>\"\",G{i},F{i})"
    ws[f"M{i}"] = f"=IF(E{i}<>\"\",E{i}*J{i},\"\")"
    ws[f"N{i}"] = f"=IF(E{i}<>\"\",E{i}*K{i},\"\")"
    ws[f"O{i}"] = f"=IF(E{i}<>\"\",E{i}*L{i},\"\")"
tr = last + 1
ws[f"A{tr}"] = "TOTAL"
ws[f"A{tr}"].font = BOLD
ws[f"E{tr}"] = f"=SUM(E{first}:E{last})"
for col in "MNO":
    ws[f"{col}{tr}"] = f"=SUM({col}{first}:{col}{last})"
    ws[f"{col}{tr}"].font = BOLD
for i in range(first, tr + 1):
    for col in ("E",):
        ws[f"{col}{i}"].number_format = "#,##0"
    for col in "FGHJKLMNO":
        ws[f"{col}{i}"].number_format = MONEY
autofit(ws, [24, 26, 10, 14, 8, 12, 12, 12, 30, 12, 12, 12, 13, 13, 13])
ws.freeze_panes = "A4"

# ---------------------------------------------------------------- Equipment
ws = wb.create_sheet("Equipment")
ws["A1"] = "FORKLIFT & WAREHOUSE EQUIPMENT (BZ International Trade LLC - 2024 stock summary; On-hand flag = May 2026 file)"
ws["A1"].font = TITLE_FONT
hdr, rows = read_csv("equipment_inventory.csv")
headers = hdr + ["Quick $/unit", "Mid $/unit", "Slow $/unit", "Line Total @ Mid (on-hand only)"]
ws.append([])
ws.append(headers)
style_header(ws, 3, len(headers))
first = 4
for r in rows:
    ws.append([r[0], r[1], r[2], r[3], num(r[4]), num(r[5]), num(r[6]), r[7], r[8]])
last = ws.max_row
for i in range(first, last + 1):
    ws[f"J{i}"] = f"=IF(F{i}<>\"\",F{i}*Assumptions!$B$7,\"\")"
    ws[f"K{i}"] = f"=IF(F{i}<>\"\",F{i}*Assumptions!$B$8,\"\")"
    ws[f"L{i}"] = f"=IF(G{i}<>\"\",G{i},IF(F{i}<>\"\",F{i}*0.95,\"\"))"
    ws[f"M{i}"] = f"=IF(H{i}=\"Yes\",E{i}*K{i},0)"
tr = last + 1
ws[f"A{tr}"] = "TOTALS"
ws[f"A{tr}"].font = BOLD
ws[f"E{tr}"] = f"=SUM(E{first}:E{last})"
ws[f"M{tr}"] = f"=SUM(M{first}:M{last})"
ws[f"M{tr}"].font = BOLD
ws[f"A{tr+1}"] = "2024 acquisition cost (all units):"
ws[f"E{tr+1}"] = f"=SUMPRODUCT(E{first}:E{last},F{first}:F{last})"
ws[f"E{tr+1}"].number_format = MONEY
ws[f"A{tr+2}"] = "Note: 'On-hand May 2026' = item appears in the May 2026 inventory file. Quantities need physical verification before offers go out."
for i in range(first, tr + 1):
    for col in "FGJKLM":
        ws[f"{col}{i}"].number_format = MONEY
autofit(ws, [18, 44, 12, 24, 10, 11, 13, 14, 34, 12, 12, 12, 18])
ws.freeze_panes = "A4"

# ---------------------------------------------------------------- Appliances
ws = wb.create_sheet("Appliances")
ws["A1"] = "APPLIANCE INVENTORY (Luxury: Gaggenau/Miele/Thermador/Bosch - new, several boxed. Mainstream: Milhurst list)"
ws["A1"].font = TITLE_FONT
hdr, rows = read_csv("appliances_inventory.csv")
headers = hdr + ["Quick $", "Mid $", "Slow $"]
ws.append([])
ws.append(headers)
style_header(ws, 3, len(headers))
first = 4
for r in rows:
    ws.append([r[0], r[1], r[2], r[3], num(r[4]), num(r[5]), num(r[6]), num(r[7]), r[8]])
last = ws.max_row
for i in range(first, last + 1):
    ws[f"J{i}"] = f"=IF(G{i}<>\"\",G{i}*Assumptions!$B$11,\"\")"
    ws[f"K{i}"] = f"=IF(G{i}<>\"\",G{i}*Assumptions!$B$12,\"\")"
    ws[f"L{i}"] = f"=IF(G{i}<>\"\",G{i}*Assumptions!$B$13,\"\")"
tr = last + 1
ws[f"A{tr}"] = "TOTALS"
ws[f"A{tr}"].font = BOLD
ws[f"F{tr}"] = f"=SUM(F{first}:F{last})"
ws[f"G{tr}"] = f"=SUM(G{first}:G{last})"
for col in "JKL":
    ws[f"{col}{tr}"] = f"=SUM({col}{first}:{col}{last})"
    ws[f"{col}{tr}"].font = BOLD
for i in range(first, tr + 1):
    for col in "FGHJKL":
        ws[f"{col}{i}"].number_format = MONEY
autofit(ws, [12, 14, 20, 42, 6, 11, 11, 13, 40, 11, 11, 11])
ws.freeze_panes = "A4"

# ---------------------------------------------------------------- Furniture
ws = wb.create_sheet("Furniture")
ws["A1"] = "FURNITURE / MATTRESSES / DISPLAYS (Moda Mobilya + Iskeceli + display units) - QUANTITIES NEED PHYSICAL COUNT"
ws["A1"].font = TITLE_FONT
hdr, rows = read_csv("furniture_inventory.csv")
ws.append([])
ws.append(hdr)
style_header(ws, 3, len(hdr))
for r in rows:
    ws.append(r)
autofit(ws, [12, 52, 20, 18, 10, 11, 8, 40])
ws.freeze_panes = "A4"

# ---------------------------------------------------------------- Mining
ws = wb.create_sheet("Mining")
ws["A1"] = "ULTIMA MINING - 13 EXPLORATION LICENSES, TURKIYE (Phase-1 complete: Matrix Geotechnologies 2024, 777 samples)"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Owner: Ultima Uluslararasi Tasimacilik Lojistik ve Dis Ticaret Ltd. Sti.  |  RISK: 2026 license & rehabilitation fees were unpaid as of June 2026 - VERIFY LICENSES ARE ALIVE BEFORE ANY MARKETING."
ws["A2"].font = Font(name=ARIAL, bold=True, color="C00000", size=10)
hdr, rows = read_csv("mining_sites.csv")
ws.append([])
ws.append(hdr)
style_header(ws, 4, len(hdr))
for r in rows:
    ws.append(r)
autofit(ws, [6, 24, 30, 56, 14, 10, 34])
ws.freeze_panes = "A5"

# ---------------------------------------------------------------- Assumptions
ws = wb.create_sheet("Assumptions")
ws["A1"] = "PRICING TIER ASSUMPTIONS (edit the blue cells; every tier column recalculates)"
ws["A1"].font = TITLE_FONT
data = [
    ("", ""),
    ("SLABS", ""),
    ("Quick-sale multiplier vs List A (bulk/one-lot exit, 7-21 days)", 0.70),
    ("Mid = List A (Jan-2025 offer sheet). Slow = List B (Jun-2025 list).", ""),
    ("EQUIPMENT", ""),
    ("Quick-sale multiplier vs unit cost (dealer/wholesale exit)", 0.55),
    ("Mid multiplier vs unit cost (direct end-user sale)", 0.75),
    ("APPLIANCES", ""),
    ("Quick multiplier vs MSRP (bulk to reseller/liquidator)", 0.45),
    ("Mid multiplier vs MSRP (contractor packages)", 0.55),
    ("Slow multiplier vs MSRP (retail one-by-one)", 0.70),
]
for row in data:
    ws.append(row)
for r in (4, 7, 8, 11, 12, 13):
    c = ws[f"B{r}"]
    c.font = Font(name=ARIAL, color="0000FF", bold=True, size=10)
    c.fill = PatternFill("solid", fgColor="FFFF00")
    c.number_format = "0%"
ws.append(("", ""))
notes = [
    "SOURCES:",
    "Slabs: 'Inventory of Slabs and Tiles.xlsx' (List A, Jan-2025, incl. 25% bulk-discount footer) + 'Inventory of Slabs Northvale June 2025.xlsx' (List B). Google Drive.",
    "Vietnam quartz actual purchase cost $481.28/slab from BZ Zoho item export ('Item (1).xlsx').",
    "Equipment: 'Forklift and Warehouse Equipment Inventory (1).xlsx' (2024 stock summary w/ costs); on-hand flags from 'Forklift and Warehouse Equipment Inventory May 2026.xls'.",
    "Appliances: 'APPLIANCES (1) (1).xlsx' (cost+MSRP) + 'APPLIANCES_updated_retail_prices.xlsx' (verified retail) + Milhurst pricing pack (2026-06-13).",
    "Mining: 'ULTIMA_Arama_Projeleri_Sunum.pdf' (13-site Phase-1 report).",
    "ALL quantities are per the source files - physical verification required before binding offers.",
    "LEGAL: No sale, transfer, or binding offer without clearance from bankruptcy counsel (Zazella & Singer) - see SALES_PLAN_30_DAYS.md, section 'Legal guardrails'.",
]
for n in notes:
    ws.append((n,))
autofit(ws, [72, 12])

# ---------------------------------------------------------------- Summary
ws = wb.create_sheet("Summary", 0)
ws["A1"] = "MASTER INVENTORY SUMMARY - all categories (USD)"
ws["A1"].font = Font(name=ARIAL, bold=True, size=14)
ws["A2"] = "Prepared 2026-08-27 from Google Drive source files. DRAFT - for internal planning and counsel review only. Not a solicitation."
ws["A2"].font = Font(name=ARIAL, italic=True, size=9)
ws.append([])
headers = ["Category", "Units", "Quick Sale (7-21d)", "Mid (30-45d)", "Slow / Retail (60-90d+)", "Status / Data gaps"]
ws.append(headers)
style_header(ws, 4, len(headers))
slab_tr = None
# find totals rows dynamically: Slabs total row = last data row + 1 computed above; recompute
slabs_last = 3 + len(read_csv("slabs_inventory.csv")[1])
eq_last = 3 + len(read_csv("equipment_inventory.csv")[1])
app_last = 3 + len(read_csv("appliances_inventory.csv")[1])
rows = [
    ("Slabs & tiles (porcelain/quartz/marble)", f"=Slabs!E{slabs_last+1}",
     f"=Slabs!M{slabs_last+1}", f"=Slabs!N{slabs_last+1}", f"=Slabs!O{slabs_last+1}",
     "Pcs per Jan-2025 list; verify current location & counts"),
    ("Forklifts & warehouse equipment", f"=Equipment!E{eq_last+1}",
     f"=SUMPRODUCT((Equipment!H4:H{eq_last}=\"Yes\")*Equipment!E4:E{eq_last}*Equipment!J4:J{eq_last})",
     f"=Equipment!M{eq_last+1}",
     f"=SUMPRODUCT((Equipment!H4:H{eq_last}=\"Yes\")*Equipment!E4:E{eq_last}*Equipment!L4:L{eq_last})",
     "On-hand rows only (May 2026 file); confirm hours/condition"),
    ("Appliances (luxury + mainstream)", f"=Appliances!F{app_last+1}",
     f"=Appliances!J{app_last+1}", f"=Appliances!K{app_last+1}", f"=Appliances!L{app_last+1}",
     "MSRP-based tiers; mainstream rows have researched asks in Notes"),
    ("Furniture / mattresses / displays", "TBD", "TBD", "TBD", "TBD",
     "Physical count required - Zoho stock shows zero for most SKUs"),
    ("Ausavina stone equipment (new stock)", "TBD", "TBD", "TBD", "TBD",
     "Use Ausavina US price list Apr-2025 as 'Slow'; count stock Sunday"),
    ("Mining licenses (13 sites, Turkiye)", 13, "n/a", "n/a", "n/a",
     "Not priced - teaser/NDA/data-room process; verify license fees paid"),
]
for r in rows:
    ws.append(r)
tr = ws.max_row + 1
ws[f"A{tr}"] = "TOTAL (priced categories)"
ws[f"A{tr}"].font = BOLD
for col in "CDE":
    ws[f"{col}{tr}"] = f"=SUM({col}5:{col}7)"
    ws[f"{col}{tr}"].font = BOLD
for i in range(5, tr + 1):
    for col in "CDE":
        ws[f"{col}{i}"].number_format = MONEY
ws.append([])
ws.append(["LEGAL: Debtor is in Chapter 11 (In re Cakmak, 26-11521-VFP, D.N.J.). No sale/transfer/binding offer of any of these assets",])
ws.append(["without prior clearance from bankruptcy counsel (Leonard S. Singer, Zazella & Singer). See SALES_PLAN_30_DAYS.md.",])
ws[f"A{ws.max_row-1}"].font = Font(name=ARIAL, bold=True, color="C00000", size=10)
ws[f"A{ws.max_row}"].font = Font(name=ARIAL, bold=True, color="C00000", size=10)
autofit(ws, [38, 10, 17, 15, 20, 52])

# global font pass for data cells that kept default font
for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None and cell.font.name != ARIAL:
                bold = cell.font.bold
                color = cell.font.color
                sz = cell.font.size if cell.font.size else 10
                cell.font = Font(name=ARIAL, bold=bold, color=color, size=sz)

OUT.parent.mkdir(exist_ok=True)
wb.save(OUT)
print(f"saved {OUT}")
